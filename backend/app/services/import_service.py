from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Literal
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import and_, delete, select
from sqlalchemy.orm import Session

from app.core.config import Settings, get_settings
from app.models import ImportJob, Product, PurchasesDaily, SalesDaily
from app.services.data_generator import DataGenerator

ImportEntityType = Literal["sales", "purchases"]
ImportJobStatus = Literal["queued", "processing", "completed", "completed_with_errors", "failed"]

TERMINAL_STATUSES: tuple[ImportJobStatus, ...] = ("completed", "completed_with_errors", "failed")
DATA_SOURCE_FILE_DEFAULT = "historical_data"
DATA_SOURCE_GENERATED = "historical_data"

SALES_REQUIRED_COLUMNS: tuple[str, ...] = (
    "date",
    "product_code",
    "volume_liters",
    "revenue_rub",
    "avg_retail_price_rub",
)
PURCHASES_REQUIRED_COLUMNS: tuple[str, ...] = (
    "date",
    "product_code",
    "volume_liters",
    "purchase_price_rub",
    "supplier_name",
    "logistics_cost_rub",
)

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("date", "sale_date", "purchase_date"),
    "product_code": ("product_code", "product", "code", "productcode"),
    "volume_liters": ("volume_liters", "volume", "liters", "volume_l"),
    "revenue_rub": ("revenue_rub", "revenue"),
    "avg_retail_price_rub": (
        "avg_retail_price_rub",
        "retail_price_rub",
        "avg_retail_price",
        "retail_price",
    ),
    "purchase_price_rub": ("purchase_price_rub", "purchase_price", "buy_price", "cost_price"),
    "supplier_name": ("supplier_name", "supplier"),
    "logistics_cost_rub": ("logistics_cost_rub", "logistics_cost", "delivery_cost"),
}


@dataclass(frozen=True)
class RowError:
    row_number: int
    code: str
    message: str
    raw: dict[str, Any] | None = None


@dataclass(frozen=True)
class PreparedSalesRow:
    row_number: int
    sale_date: date
    product_id: UUID
    product_code: str
    volume_liters: Decimal
    revenue_rub: Decimal
    avg_retail_price_rub: Decimal


@dataclass(frozen=True)
class PreparedPurchaseRow:
    row_number: int
    purchase_date: date
    product_id: UUID
    product_code: str
    volume_liters: Decimal
    purchase_price_rub: Decimal
    logistics_cost_rub: Decimal
    supplier_name: str | None
    total_cost_rub: Decimal


@dataclass(frozen=True)
class GenerateDemoPayload:
    start_date: date
    end_date: date
    products: list[str]
    seed: int
    replace_existing: bool


class ImportService:
    def __init__(self, session: Session, settings: Settings | None = None) -> None:
        self._session = session
        self._settings = settings or get_settings()

    def create_job(
        self,
        *,
        entity_type: str,
        source_type: str,
        file_name: str | None,
        started_by: UUID,
    ) -> ImportJob:
        job = ImportJob(
            entity_type=entity_type,
            source_type=source_type,
            file_name=file_name,
            status="queued",
            started_by=started_by,
        )
        self._session.add(job)
        self._session.commit()
        self._session.refresh(job)
        return job

    def list_jobs(
        self,
        *,
        entity_type: str | None = None,
        status: str | None = None,
        limit: int = 20,
    ) -> list[ImportJob]:
        query = select(ImportJob).order_by(ImportJob.started_at.desc()).limit(limit)
        if entity_type:
            query = query.where(ImportJob.entity_type == entity_type)
        if status:
            query = query.where(ImportJob.status == status)
        return list(self._session.scalars(query))

    def get_job(self, *, job_id: UUID) -> ImportJob | None:
        return self._session.get(ImportJob, job_id)

    def process_file_job(
        self,
        *,
        job_id: UUID,
        entity_type: ImportEntityType,
        file_name: str,
        file_bytes: bytes,
        source_name: str | None,
    ) -> None:
        job = self._require_job(job_id=job_id)
        try:
            self._set_job_processing(job)
            rows = self._read_rows(file_name=file_name, file_bytes=file_bytes)
            if not rows:
                error = RowError(row_number=0, code="empty_file", message="Файл не содержит данных")
                self._finalize_job(job, rows_total=0, rows_success=0, errors=[error])
                return

            if entity_type == "sales":
                success_rows, errors = self._prepare_sales_rows(rows=rows)
                self._persist_sales_rows(job=job, rows=success_rows, source_name=source_name)
                self._finalize_job(
                    job,
                    rows_total=len(rows),
                    rows_success=len(success_rows),
                    errors=errors,
                )
                return

            success_rows, errors = self._prepare_purchase_rows(rows=rows)
            self._persist_purchase_rows(job=job, rows=success_rows, source_name=source_name)
            self._finalize_job(
                job,
                rows_total=len(rows),
                rows_success=len(success_rows),
                errors=errors,
            )
        except Exception as exc:  # pragma: no cover
            self._fail_job(job, reason=str(exc))

    def process_generate_demo_job(self, *, job_id: UUID, payload: GenerateDemoPayload) -> None:
        job = self._require_job(job_id=job_id)
        try:
            self._set_job_processing(job)
            products = self._load_products_by_codes(payload.products)
            unknown_codes = sorted(set(payload.products) - set(products.keys()))
            if unknown_codes:
                error = RowError(
                    row_number=0,
                    code="unknown_product_code",
                    message=f"Неизвестные product_code: {', '.join(unknown_codes)}",
                )
                self._finalize_job(job, rows_total=0, rows_success=0, errors=[error])
                return

            if payload.replace_existing:
                self._delete_existing_history(
                    start_date=payload.start_date,
                    end_date=payload.end_date,
                    product_ids=[product.id for product in products.values()],
                )

            generator = DataGenerator(seed=payload.seed)
            dataset = generator.generate(
                start_date=payload.start_date,
                end_date=payload.end_date,
                product_codes=payload.products,
            )

            # Map generated rows to ORM objects
            sales_orm = self._map_sales_to_orm(
                rows=dataset.sales, products=products, batch_id=job.id,
            )
            purchase_orm = self._map_purchases_to_orm(
                rows=dataset.purchases, products=products, batch_id=job.id,
            )

            # Batched insert (500 per chunk)
            for i in range(0, len(sales_orm), 500):
                self._session.add_all(sales_orm[i : i + 500])
                self._session.flush()
            for i in range(0, len(purchase_orm), 500):
                self._session.add_all(purchase_orm[i : i + 500])
                self._session.flush()

            generated_rows_total = len(payload.products) * (
                (payload.end_date - payload.start_date).days + 1
            )
            self._finalize_job(
                job,
                rows_total=generated_rows_total,
                rows_success=len(sales_orm),
                errors=[],
            )
        except Exception as exc:  # pragma: no cover
            self._fail_job(job, reason=str(exc))

    def _require_job(self, *, job_id: UUID) -> ImportJob:
        job = self._session.get(ImportJob, job_id)
        if job is None:
            raise ValueError(f"Import job {job_id} not found")
        return job

    def _set_job_processing(self, job: ImportJob) -> None:
        job.status = "processing"
        self._session.commit()

    def _finalize_job(
        self,
        job: ImportJob,
        *,
        rows_total: int,
        rows_success: int,
        errors: list[RowError],
    ) -> None:
        rows_failed = len(errors)
        status = self._resolve_terminal_status(
            rows_total=rows_total,
            rows_success=rows_success,
            rows_failed=rows_failed,
        )
        error_report_path = self._write_error_report(job_id=job.id, errors=errors)

        job.rows_total = rows_total
        job.rows_success = rows_success
        job.rows_failed = rows_failed
        job.error_report_path = error_report_path
        job.status = status
        job.finished_at = datetime.now(UTC)
        self._session.commit()

    @staticmethod
    def _resolve_terminal_status(
        *,
        rows_total: int,
        rows_success: int,
        rows_failed: int,
    ) -> ImportJobStatus:
        if rows_success == rows_total and rows_failed == 0 and rows_total > 0:
            return "completed"
        if rows_success > 0 and rows_failed > 0:
            return "completed_with_errors"
        return "failed"

    def _fail_job(self, job: ImportJob, *, reason: str) -> None:
        report_path = self._write_error_report(
            job_id=job.id,
            errors=[RowError(row_number=0, code="internal_error", message=reason)],
        )
        job.status = "failed"
        job.error_report_path = report_path
        job.finished_at = datetime.now(UTC)
        self._session.commit()

    def _prepare_sales_rows(
        self,
        *,
        rows: list[dict[str, Any]],
    ) -> tuple[list[PreparedSalesRow], list[RowError]]:
        required = SALES_REQUIRED_COLUMNS
        canonical_rows, header_errors = self._canonicalize_rows(
            rows=rows,
            required_columns=required,
        )
        if header_errors:
            return [], header_errors

        product_codes = sorted({str(item["product_code"]).upper() for item in canonical_rows})
        products = self._load_products_by_codes(product_codes)
        errors: list[RowError] = []
        seen_keys: set[tuple[date, str]] = set()
        prepared: list[PreparedSalesRow] = []

        for row in canonical_rows:
            row_number = int(row["_row_number"])
            try:
                row_date = self._parse_date(row["date"])
                code = str(row["product_code"]).upper().strip()
                product = products.get(code)
                if product is None:
                    errors.append(
                        RowError(
                            row_number=row_number,
                            code="unknown_product_code",
                            message=f"Неизвестный product_code: {code}",
                            raw=row,
                        )
                    )
                    continue
                duplicate_key = (row_date, code)
                if duplicate_key in seen_keys:
                    errors.append(
                        RowError(
                            row_number=row_number,
                            code="duplicate_in_file",
                            message=f"Дубликат строки для {code} на {row_date}",
                            raw=row,
                        )
                    )
                    continue
                seen_keys.add(duplicate_key)

                volume_liters = self._parse_decimal(row["volume_liters"], positive=True)
                revenue_rub = self._parse_decimal(row["revenue_rub"], positive=True)
                avg_retail_price_rub = self._parse_decimal(
                    row["avg_retail_price_rub"],
                    positive=True,
                )
                prepared.append(
                    PreparedSalesRow(
                        row_number=row_number,
                        sale_date=row_date,
                        product_id=product.id,
                        product_code=code,
                        volume_liters=volume_liters,
                        revenue_rub=revenue_rub,
                        avg_retail_price_rub=avg_retail_price_rub,
                    )
                )
            except ValueError as exc:
                errors.append(
                    RowError(
                        row_number=row_number,
                        code="validation_error",
                        message=str(exc),
                        raw=row,
                    )
                )

        dedupe = self._filter_existing_sales_duplicates(prepared_rows=prepared)
        errors.extend(dedupe["errors"])
        return dedupe["rows"], errors

    def _prepare_purchase_rows(
        self,
        *,
        rows: list[dict[str, Any]],
    ) -> tuple[list[PreparedPurchaseRow], list[RowError]]:
        required = PURCHASES_REQUIRED_COLUMNS
        canonical_rows, header_errors = self._canonicalize_rows(
            rows=rows,
            required_columns=required,
        )
        if header_errors:
            return [], header_errors

        product_codes = sorted({str(item["product_code"]).upper() for item in canonical_rows})
        products = self._load_products_by_codes(product_codes)
        errors: list[RowError] = []
        seen_keys: set[tuple[date, str]] = set()
        prepared: list[PreparedPurchaseRow] = []
        today = datetime.now(UTC).date()

        for row in canonical_rows:
            row_number = int(row["_row_number"])
            try:
                row_date = self._parse_date(row["date"])
                if row_date > today:
                    raise ValueError("Дата закупки не может быть позже текущего дня")

                code = str(row["product_code"]).upper().strip()
                product = products.get(code)
                if product is None:
                    errors.append(
                        RowError(
                            row_number=row_number,
                            code="unknown_product_code",
                            message=f"Неизвестный product_code: {code}",
                            raw=row,
                        )
                    )
                    continue
                duplicate_key = (row_date, code)
                if duplicate_key in seen_keys:
                    errors.append(
                        RowError(
                            row_number=row_number,
                            code="duplicate_in_file",
                            message=f"Дубликат строки для {code} на {row_date}",
                            raw=row,
                        )
                    )
                    continue
                seen_keys.add(duplicate_key)

                volume_liters = self._parse_decimal(row["volume_liters"], positive=True)
                purchase_price_rub = self._parse_decimal(
                    row["purchase_price_rub"],
                    non_negative=True,
                )
                logistics_cost_rub = self._parse_decimal(
                    row["logistics_cost_rub"],
                    non_negative=True,
                )
                supplier_name = str(row["supplier_name"]).strip() or None
                total_cost_rub = (volume_liters * purchase_price_rub) + logistics_cost_rub
                prepared.append(
                    PreparedPurchaseRow(
                        row_number=row_number,
                        purchase_date=row_date,
                        product_id=product.id,
                        product_code=code,
                        volume_liters=volume_liters,
                        purchase_price_rub=purchase_price_rub,
                        logistics_cost_rub=logistics_cost_rub,
                        supplier_name=supplier_name,
                        total_cost_rub=total_cost_rub,
                    )
                )
            except ValueError as exc:
                errors.append(
                    RowError(
                        row_number=row_number,
                        code="validation_error",
                        message=str(exc),
                        raw=row,
                    )
                )

        dedupe = self._filter_existing_purchase_duplicates(prepared_rows=prepared)
        errors.extend(dedupe["errors"])
        return dedupe["rows"], errors

    def _filter_existing_sales_duplicates(
        self,
        *,
        prepared_rows: list[PreparedSalesRow],
    ) -> dict[str, list[Any]]:
        if not prepared_rows:
            return {"rows": [], "errors": []}
        dates = sorted({row.sale_date for row in prepared_rows})
        product_ids = list({row.product_id for row in prepared_rows})
        existing = set(
            self._session.execute(
                select(SalesDaily.sale_date, SalesDaily.product_id).where(
                    and_(
                        SalesDaily.sale_date >= dates[0],
                        SalesDaily.sale_date <= dates[-1],
                        SalesDaily.product_id.in_(product_ids),
                    )
                )
            ).all()
        )

        rows: list[PreparedSalesRow] = []
        errors: list[RowError] = []
        for item in prepared_rows:
            key = (item.sale_date, item.product_id)
            if key in existing:
                errors.append(
                        RowError(
                            row_number=item.row_number,
                            code="duplicate_existing",
                            message=(
                                f"Данные уже существуют для {item.product_code}"
                                f" на {item.sale_date}"
                            ),
                        )
                    )
                continue
            rows.append(item)
        return {"rows": rows, "errors": errors}

    def _filter_existing_purchase_duplicates(
        self,
        *,
        prepared_rows: list[PreparedPurchaseRow],
    ) -> dict[str, list[Any]]:
        if not prepared_rows:
            return {"rows": [], "errors": []}
        dates = sorted({row.purchase_date for row in prepared_rows})
        product_ids = list({row.product_id for row in prepared_rows})
        existing = set(
            self._session.execute(
                select(PurchasesDaily.purchase_date, PurchasesDaily.product_id).where(
                    and_(
                        PurchasesDaily.purchase_date >= dates[0],
                        PurchasesDaily.purchase_date <= dates[-1],
                        PurchasesDaily.product_id.in_(product_ids),
                    )
                )
            ).all()
        )

        rows: list[PreparedPurchaseRow] = []
        errors: list[RowError] = []
        for item in prepared_rows:
            key = (item.purchase_date, item.product_id)
            if key in existing:
                errors.append(
                        RowError(
                            row_number=item.row_number,
                            code="duplicate_existing",
                            message=(
                                f"Данные уже существуют для {item.product_code}"
                                f" на {item.purchase_date}"
                            ),
                        )
                    )
                continue
            rows.append(item)
        return {"rows": rows, "errors": errors}

    def _persist_sales_rows(
        self,
        *,
        job: ImportJob,
        rows: list[PreparedSalesRow],
        source_name: str | None,
    ) -> None:
        if not rows:
            return
        data_source = self._build_data_source(source_name)
        records = [
            SalesDaily(
                sale_date=row.sale_date,
                product_id=row.product_id,
                volume_liters=row.volume_liters,
                revenue_rub=row.revenue_rub,
                avg_retail_price_rub=row.avg_retail_price_rub,
                data_source=data_source,
                source_batch_id=job.id,
            )
            for row in rows
        ]
        self._session.add_all(records)
        self._session.flush()

    def _persist_purchase_rows(
        self,
        *,
        job: ImportJob,
        rows: list[PreparedPurchaseRow],
        source_name: str | None,
    ) -> None:
        if not rows:
            return
        data_source = self._build_data_source(source_name)
        records = [
            PurchasesDaily(
                purchase_date=row.purchase_date,
                product_id=row.product_id,
                volume_liters=row.volume_liters,
                purchase_price_rub=row.purchase_price_rub,
                logistics_cost_rub=row.logistics_cost_rub,
                supplier_name=row.supplier_name,
                total_cost_rub=row.total_cost_rub,
                data_source=data_source,
                source_batch_id=job.id,
            )
            for row in rows
        ]
        self._session.add_all(records)
        self._session.flush()

    @staticmethod
    def _build_data_source(source_name: str | None) -> str:
        if source_name and source_name.strip():
            return source_name.strip()[:32]
        return DATA_SOURCE_FILE_DEFAULT

    def _map_sales_to_orm(
        self,
        *,
        rows: list,
        products: dict[str, Product],
        batch_id: UUID,
    ) -> list[SalesDaily]:
        return [
            SalesDaily(
                sale_date=row.sale_date,
                product_id=products[row.product_code].id,
                volume_liters=row.volume_liters,
                revenue_rub=row.revenue_rub,
                avg_retail_price_rub=row.avg_retail_price_rub,
                data_source=DATA_SOURCE_GENERATED,
                source_batch_id=batch_id,
            )
            for row in rows
        ]

    def _map_purchases_to_orm(
        self,
        *,
        rows: list,
        products: dict[str, Product],
        batch_id: UUID,
    ) -> list[PurchasesDaily]:
        return [
            PurchasesDaily(
                purchase_date=row.purchase_date,
                product_id=products[row.product_code].id,
                volume_liters=row.volume_liters,
                purchase_price_rub=row.purchase_price_rub,
                logistics_cost_rub=row.logistics_cost_rub,
                supplier_name=row.supplier_name,
                total_cost_rub=row.total_cost_rub,
                data_source=DATA_SOURCE_GENERATED,
                source_batch_id=batch_id,
            )
            for row in rows
        ]

    def _delete_existing_history(
        self,
        *,
        start_date: date,
        end_date: date,
        product_ids: list[UUID],
    ) -> None:
        if not product_ids:
            return
        self._session.execute(
            delete(SalesDaily).where(
                and_(
                    SalesDaily.sale_date >= start_date,
                    SalesDaily.sale_date <= end_date,
                    SalesDaily.product_id.in_(product_ids),
                )
            )
        )
        self._session.execute(
            delete(PurchasesDaily).where(
                and_(
                    PurchasesDaily.purchase_date >= start_date,
                    PurchasesDaily.purchase_date <= end_date,
                    PurchasesDaily.product_id.in_(product_ids),
                )
            )
        )
        self._session.flush()

    def _load_products_by_codes(self, product_codes: list[str]) -> dict[str, Product]:
        if not product_codes:
            return {}
        normalized = [code.upper().strip() for code in product_codes]
        rows = self._session.scalars(select(Product).where(Product.code.in_(normalized)))
        return {product.code.upper(): product for product in rows}

    @staticmethod
    def _parse_date(value: Any) -> date:
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        text = str(value).strip()
        if not text:
            raise ValueError("Дата не заполнена")
        for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
        raise ValueError(f"Некорректный формат даты: {text}")

    @staticmethod
    def _parse_decimal(
        value: Any,
        *,
        positive: bool = False,
        non_negative: bool = False,
    ) -> Decimal:
        text = str(value).strip().replace(",", ".")
        if not text:
            raise ValueError("Значение не заполнено")
        try:
            number = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError(f"Некорректное числовое значение: {value}") from exc
        if positive and number <= 0:
            raise ValueError("Значение должно быть больше 0")
        if non_negative and number < 0:
            raise ValueError("Значение не может быть отрицательным")
        return number

    def _canonicalize_rows(
        self,
        *,
        rows: list[dict[str, Any]],
        required_columns: tuple[str, ...],
    ) -> tuple[list[dict[str, Any]], list[RowError]]:
        if not rows:
            return [], []
        normalized_headers = {self._normalize_header(key): key for key in rows[0].keys()}
        missing = [
            column
            for column in required_columns
            if not any(alias in normalized_headers for alias in COLUMN_ALIASES[column])
        ]
        if missing:
            return (
                [],
                [
                    RowError(
                        row_number=0,
                        code="missing_columns",
                        message=f"Отсутствуют обязательные колонки: {', '.join(missing)}",
                    )
                ],
            )

        canonical_rows: list[dict[str, Any]] = []
        for row in rows:
            row_number = int(row.get("_row_number", 0))
            canonical: dict[str, Any] = {"_row_number": row_number}
            for column in required_columns:
                value = None
                for alias in COLUMN_ALIASES[column]:
                    source_key = normalized_headers.get(alias)
                    if source_key is None:
                        continue
                    value = row.get(source_key)
                    break
                canonical[column] = value
            canonical_rows.append(canonical)
        return canonical_rows, []

    @staticmethod
    def _normalize_header(value: str) -> str:
        return (
            value.strip()
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
            .replace(".", "_")
            .replace("/", "_")
        )

    def _read_rows(self, *, file_name: str, file_bytes: bytes) -> list[dict[str, Any]]:
        extension = Path(file_name).suffix.lower()
        if extension == ".csv":
            return self._read_csv(file_bytes=file_bytes)
        if extension == ".xlsx":
            return self._read_xlsx(file_bytes=file_bytes)
        raise ValueError(f"Неподдерживаемый формат файла: {extension}")

    def _read_csv(self, *, file_bytes: bytes) -> list[dict[str, Any]]:
        decoded = self._decode_csv(file_bytes=file_bytes)
        if not decoded.strip():
            return []
        sample = decoded[:1024]
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
        reader = csv.DictReader(StringIO(decoded), delimiter=delimiter)
        rows: list[dict[str, Any]] = []
        for index, row in enumerate(reader, start=2):
            clean_row = {str(key).strip(): value for key, value in row.items() if key is not None}
            clean_row["_row_number"] = index
            rows.append(clean_row)
        return rows

    @staticmethod
    def _decode_csv(*, file_bytes: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError("Не удалось прочитать CSV в поддерживаемой кодировке")

    @staticmethod
    def _read_xlsx(*, file_bytes: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: list[dict[str, Any]] = []
        for index, values in enumerate(rows_iter, start=2):
            row: dict[str, Any] = {"_row_number": index}
            for column_index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[column_index] if column_index < len(values) else None
            rows.append(row)
        return rows



    def _write_error_report(self, *, job_id: UUID, errors: list[RowError]) -> str | None:
        if not errors:
            return None
        reports_dir = self._resolve_reports_dir()
        reports_dir.mkdir(parents=True, exist_ok=True)
        report_path = reports_dir / f"import_job_{job_id}.json"
        payload = {
            "job_id": str(job_id),
            "generated_at": datetime.now(UTC).isoformat(),
            "errors": [
                {
                    "row_number": item.row_number,
                    "code": item.code,
                    "message": item.message,
                    "raw": item.raw,
                }
                for item in errors
            ],
        }
        report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return str(report_path)

    def _resolve_reports_dir(self) -> Path:
        preferred = Path(self._settings.model_artifacts_dir) / "import_reports"
        try:
            preferred.mkdir(parents=True, exist_ok=True)
            return preferred
        except OSError:
            fallback = Path(".artifacts") / "import_reports"
            fallback.mkdir(parents=True, exist_ok=True)
            return fallback
